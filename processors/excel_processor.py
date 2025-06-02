import pandas as pd
import openpyxl
from typing import Dict, Any, List
from .base_processor import BaseProcessor

class ExcelProcessor(BaseProcessor):
    """Process Excel and CSV files"""
    
    def process(self, filepath: str) -> Dict[str, Any]:
        """Extract data from spreadsheet files"""
        result = {
            'text': '',
            'sheets': [],
            'metadata': {}
        }
        
        try:
            file_extension = filepath.split('.')[-1].lower()
            
            if file_extension == 'csv':
                result = self._process_csv(filepath)
            else:
                result = self._process_excel(filepath)
                
        except Exception as e:
            result['error'] = str(e)
            result['text'] = f"Error processing spreadsheet: {str(e)}"
        
        return result
    
    def _process_csv(self, filepath: str) -> Dict[str, Any]:
        """Process CSV file"""
        result = {
            'text': '',
            'sheets': [],
            'metadata': {'file_type': 'CSV'}
        }
        
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(filepath, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                df = pd.read_csv(filepath, encoding='utf-8', errors='ignore')
            
            sheet_data = {
                'name': 'Sheet1',
                'data': df.to_dict('records'),
                'columns': df.columns.tolist(),
                'shape': df.shape
            }
            
            result['sheets'].append(sheet_data)
            result['metadata']['rows'] = df.shape[0]
            result['metadata']['columns'] = df.shape[1]
            
            # Convert to text
            text_parts = [f"# CSV Data\n\n**Rows:** {df.shape[0]}\n**Columns:** {df.shape[1]}\n\n"]
            text_parts.append("## Columns\n")
            text_parts.extend([f"- {col}" for col in df.columns])
            text_parts.append("\n\n## Data\n\n")
            text_parts.append(df.to_markdown(index=False) if hasattr(df, 'to_markdown') else df.to_string())
            
            result['text'] = '\n'.join(text_parts)
            
        except Exception as e:
            result['error'] = str(e)
            result['text'] = f"Error processing CSV: {str(e)}"
        
        return result
    
    def _process_excel(self, filepath: str) -> Dict[str, Any]:
        """Process Excel file"""
        result = {
            'text': '',
            'sheets': [],
            'metadata': {}
        }
        
        try:
            # Get workbook info
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            result['metadata'] = {
                'file_type': 'Excel',
                'sheet_names': wb.sheetnames,
                'sheet_count': len(wb.sheetnames)
            }
            
            text_parts = [f"# Excel Workbook\n\n**Sheets:** {len(wb.sheetnames)}\n\n"]
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                try:
                    df = pd.read_excel(filepath, sheet_name=sheet_name)
                    
                    sheet_data = {
                        'name': sheet_name,
                        'data': df.to_dict('records'),
                        'columns': df.columns.tolist(),
                        'shape': df.shape
                    }
                    
                    result['sheets'].append(sheet_data)
                    
                    # Add to text
                    text_parts.append(f"## Sheet: {sheet_name}\n\n")
                    text_parts.append(f"**Rows:** {df.shape[0]}\n**Columns:** {df.shape[1]}\n\n")
                    
                    if not df.empty:
                        text_parts.append("### Columns\n")
                        text_parts.extend([f"- {col}" for col in df.columns])
                        text_parts.append("\n\n### Data\n\n")
                        text_parts.append(df.to_markdown(index=False) if hasattr(df, 'to_markdown') else df.to_string())
                        text_parts.append("\n\n")
                    
                except Exception as e:
                    text_parts.append(f"Error processing sheet '{sheet_name}': {str(e)}\n\n")
            
            result['text'] = '\n'.join(text_parts)
            
        except Exception as e:
            result['error'] = str(e)
            result['text'] = f"Error processing Excel file: {str(e)}"
        
        return result
    
    def to_markdown(self, content: Dict[str, Any]) -> str:
        """Convert spreadsheet content to markdown"""
        if isinstance(content, str):
            return content
        
        if content.get('text'):
            return content['text']
        
        # Fallback conversion
        markdown = "# Spreadsheet Data\n\n"
        
        if content.get('metadata'):
            metadata = content['metadata']
            for key, value in metadata.items():
                markdown += f"**{key.title()}:** {value}\n\n"
        
        if content.get('sheets'):
            for sheet in content['sheets']:
                markdown += f"## {sheet['name']}\n\n"
                markdown += f"Rows: {sheet['shape'][0]}, Columns: {sheet['shape'][1]}\n\n"
                
                # Sample data
                if sheet['data'] and len(sheet['data']) > 0:
                    markdown += "### Sample Data (first 5 rows)\n\n"
                    sample_data = sheet['data'][:5]
                    
                    if sample_data:
                        headers = list(sample_data[0].keys())
                        markdown += "| " + " | ".join(headers) + " |\n"
                        markdown += "| " + " | ".join(["---"] * len(headers)) + " |\n"
                        
                        for row in sample_data:
                            values = [str(row.get(h, '')) for h in headers]
                            markdown += "| " + " | ".join(values) + " |\n"
                    
                    markdown += "\n\n"
        
        return markdown
